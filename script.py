from hashlib import new
from openpyxl.utils.cell import rows_from_range
import requests
import openpyxl
import json
import os
import time
import sys

from otp import SENDSMS


ip_address = input('Ip address:')

input_file_name =  input('Oxuyacagi faylin tam adini daxil edin:') or 'files/operations.xlsx' 
output_file_name = input('Neticeleri yazacagi faylin tam adinin daxil edin:') or 'results.xlsx'


try:
    wb = openpyxl.load_workbook(os.path.abspath(input_file_name))
    wb2 = openpyxl.load_workbook(os.path.abspath(output_file_name))

except Exception as e:
    print(e , '\n' , 'Fayllarin acilmasinda problem yasandi eger varsa fayli excelde duzeldib elave edin yoxdursa bura elave edin')
    sys.exit()





nums_list =  ['994552858577' , '994512059558']


ws = wb.active
ws2 = wb2.active

settings = json.loads(open('settings.json').read())


def send_sms_on_error(message):
    for num in nums_list:
        SENDSMS.sendOTP(num , message)


    
def make_data(token_request : dict , token_response : dict , *args, **kwargs) -> json:
    new_data = dict()
    new_data['requestData']  = dict()
    new_data['requestData']['tokenData'] = {'operationId':'createDocument'}


    new_data['requestData']['tokenData'].update(token_request)
    operation_data = {
        'firstOperationAtUtc':"",
        'lastOperationAtUtc':"",
        'parentDocument':token_response['data']['document_id'],
        'refund_document_number':token_response['data']['document_number'],
        'refund_short_document_id':token_response['data']['short_document_id']
        }
    new_data['requestData']['tokenData']['parameters']['data'].update(operation_data)
    new_data['requestData']['tokenData']['parameters']['doc_type']  = 'money_back'
    new_data['requestData']['checkData'] = {'check_type':100}
    new_data = json.dumps(new_data)
    return new_data


print(settings['last_row'])
for row in  ws.iter_rows(min_row=settings['last_row'], max_row=ws.max_row):
    new_data = dict()
    if row[13].value == "":
        settings['last_row'] = row[0].row
        open('settings.json','w').write(json.dumps(settings))
        wb2.save(output_file_name)
        sys.exit('Finished')

    try:
        token_request = json.loads(row[13].value)
        token_response = json.loads(row[14].value)
        print('token request' ,token_request)
    except Exception as e:
        sys.exit('Datani oxuyarken xeta bas verdi')

    new_data = make_data(token_request , token_response)
    print('new_data',new_data)
    try:
        response = requests.post(f'http://{ip_address}:8989' , new_data)

    except Exception as e:
        message = f"{row[0].row} xeta cixdi iterasiya dayandirildi"
        # send_sms_on_error(message)
        settings['last_row'] = row[0].row
        settings['status'] = 'Iterasiya qirildi'
        open('settings.json','w').write(json.dumps(settings))
        wb2.save(output_file_name)
        sys.exit('Iterasiya qirildi')
    
    print(response)
    response_data = dict(response.json())

    request_delivery_status = "Göndərilib"
    if response_data['code'] != 0:
        request_delivery_status =  "Xeta cixdi"

    ws2[row[0].coordinate] = str(new_data)
    ws2[row[1].coordinate] = request_delivery_status
    ws2[row[2].coordinate] = response.text



    
settings['last_row'] = row[0].row
settings['status']  = "Finished"
open('settings.json','w').write(json.dumps(settings))
wb2.save(output_file_name)
sys.exit('Finished')





    





