from hashlib import new
from openpyxl.utils.cell import rows_from_range
import requests
import openpyxl
import json
import os
import time
import sys

from otp import SENDSMS


ip_address = input('Ip address:')

input_file_name =  input('Oxuyacagi faylin tam adini daxil edin:') or 'files/operations.xlsx' 
output_file_name = input('Neticeleri yazacagi faylin tam adinin daxil edin:') or 'results.xlsx'


try:
    wb = openpyxl.load_workbook(os.path.abspath(input_file_name))
    wb2 = openpyxl.load_workbook(os.path.abspath(output_file_name))

except Exception as e:
    print(e , '\n' , 'Fayllarin acilmasinda(error yuxaridaki setirde) problem yasandi eger varsa fayli excelde duzeldib elave edin yoxdursa bura elave edin')
    sys.exit()


nums_list =  ['994552858577' , '994512059558']


ws = wb.active
ws2 = wb2.active
settings = json.loads(open('settings.json').read())
row_count = settings['last_row']




# save setting file
def save_setting(settings,file_name='settings.json' , status = 'Finished'):
    global row_count
    settings['last_row'] = row_count
    settings['status']  = status
    try:
        open('settings.json','w').write(json.dumps(settings))
    except Exception as e:
        wb2.save(output_file_name)
        sys.exit(f'settingi yazarken xeta cixdi sonuncu setir {row_count}')




# return response and handling error while on requests
def send_request(ip_address ,data,port=8989, *args, **kwargs):
    global row_count

    try:
        response = requests.post(f'http://{ip_address}:{port}' , data)

    except Exception as e:
        message = f"Request gonderilken xeta cixdi.{row_count} row-da xeta cixdi iterasiya dayandirildi"
        save_setting(settings,status='Iterasiya qirildi')
        wb2.save(output_file_name)
        sys.exit(message)
    
    return response





def send_sms_on_error(message):
    for num in nums_list:
        SENDSMS.sendOTP(num , message)


# for make data from token request and token response from excel    
def make_data(token_request : dict , token_response : dict , *args, **kwargs) -> json:
    new_data = dict()
    new_data['requestData']  = dict()
    new_data['requestData']['tokenData'] = {'operationId':'createDocument'}


    new_data['requestData']['tokenData'].update(token_request['inputNewJson'])
    operation_data = {
        'firstOperationAtUtc':"",
        'lastOperationAtUtc':"",
        'parentDocument':token_response['data']['document_id'],
        'refund_document_number':token_response['data']['document_number'],
        'refund_short_document_id':token_response['data']['short_document_id']
        }
    new_data['requestData']['tokenData']['parameters']['data'].update(operation_data)
    new_data['requestData']['tokenData']['parameters']['doc_type']  = 'money_back'
    new_data['requestData']['checkData'] = {'check_type':100}
    new_data = json.dumps(new_data)
    return new_data




def main():
    global row_count

    
 

    for row in  ws.iter_rows(min_row=row_count, max_row=ws.max_row):
        new_data = dict()
        if row[13].value is None:
            wb2.save(output_file_name)
            save_setting(settings)
            sys.exit('Finished')

        try:
            token_request = json.loads(row[13].value)
            token_response = json.loads(row[14].value)
        except Exception as e:
            wb2.save(output_file_name)
            sys.exit('Datani oxuyarken xeta bas verdi')

        new_data = make_data(token_request , token_response)
        
        response = send_request(ip_address ,new_data)
        response_data = dict(response.json())

        request_delivery_status = "Göndərilib"
        if response_data['code'] != 0:
            request_delivery_status =  "Xeta cixdi"

        ws2[row[0].coordinate] = str(new_data)
        ws2[row[1].coordinate] = request_delivery_status
        ws2[row[2].coordinate] = response.text
        wb2.save(output_file_name)
        print(response ,'saved' , row_count)

        # Her 100 luy rowu kecende ayri faylda save elesin 

        # if row_count % 100 == 0:
        #     print(row_count)
        #     new_file_name = f"results[{row_count-100}-{row_count}]"
        #     wb2.save(new_file_name)
        
        row_count+=1
        time.sleep(5)    
        

    save_setting(settings)
    
    wb2.save(output_file_name)
    sys.exit('Finished')


if __name__  == '__main__':
    

    try:
        main()
    except KeyboardInterrupt:
        wb2.save(output_file_name)
        wb.save()
        save_setting(settings)
        sys.exit('Results and setting saved when keyboard interruption')





    





from hashlib import new
from openpyxl.utils.cell import rows_from_range
import requests
import openpyxl
import json
import os
import time
import sys

from otp import SENDSMS


ip_address = input('Ip address:')

input_file_name =  input('Oxuyacagi faylin tam adini daxil edin:') or 'files/operations.xlsx' 
output_file_name = input('Neticeleri yazacagi faylin tam adinin daxil edin:') or 'results.xlsx'


try:
    wb = openpyxl.load_workbook(os.path.abspath(input_file_name))
    wb2 = openpyxl.load_workbook(os.path.abspath(output_file_name))

except Exception as e:
    print(e , '\n' , 'Fayllarin acilmasinda(error yuxaridaki setirde) problem yasandi eger varsa fayli excelde duzeldib elave edin yoxdursa bura elave edin')
    sys.exit()


nums_list =  ['994552858577' , '994512059558']


ws = wb.active
ws2 = wb2.active
settings = json.loads(open('settings.json').read())
row_count = settings['last_row']




# save setting file
def save_setting(settings,file_name='settings.json' , status = 'Finished'):
    global row_count
    settings['last_row'] = row_count
    settings['status']  = status
    try:
        open('settings.json','w').write(json.dumps(settings))
    except Exception as e:
        wb2.save(output_file_name)
        sys.exit(f'settingi yazarken xeta cixdi sonuncu setir {row_count}')




# return response and handling error while on requests
def send_request(ip_address ,data,port=8989, *args, **kwargs):
    global row_count

    try:
        response = requests.post(f'http://{ip_address}:{port}' , data)

    except Exception as e:
        message = f"Request gonderilken xeta cixdi.{row_count} row-da xeta cixdi iterasiya dayandirildi"
        save_setting(settings,status='Iterasiya qirildi')
        wb2.save(output_file_name)
        sys.exit(message)
    
    return response





def send_sms_on_error(message):
    for num in nums_list:
        SENDSMS.sendOTP(num , message)


# for make data from token request and token response from excel    
def make_data(token_request : dict , token_response : dict , *args, **kwargs) -> json:
    new_data = dict()
    new_data['requestData']  = dict()
    new_data['requestData']['tokenData'] = {'operationId':'createDocument'}


    new_data['requestData']['tokenData'].update(token_request)
    operation_data = {
        'firstOperationAtUtc':"",
        'lastOperationAtUtc':"",
        'parentDocument':token_response['data']['document_id'],
        'refund_document_number':token_response['data']['document_number'],
        'refund_short_document_id':token_response['data']['short_document_id']
        }
    new_data['requestData']['tokenData']['parameters']['data'].update(operation_data)
    new_data['requestData']['tokenData']['parameters']['doc_type']  = 'money_back'
    new_data['requestData']['checkData'] = {'check_type':100}
    new_data = json.dumps(new_data)
    return new_data




def main():
    global row_count

    
 

    for row in  ws.iter_rows(min_row=row_count, max_row=ws.max_row):
        new_data = dict()
        if row[13].value is None:
            wb2.save(output_file_name)
            save_setting(settings)
            sys.exit('Finished')

        try:
            token_request = json.loads(row[13].value)
            token_response = json.loads(row[14].value)
        except Exception as e:
            wb2.save(output_file_name)
            sys.exit('Datani oxuyarken xeta bas verdi')

        new_data = make_data(token_request , token_response)
        
        response = send_request(ip_address ,new_data)
        response_data = dict(response.json())

        request_delivery_status = "Göndərilib"
        if response_data['code'] != 0:
            request_delivery_status =  "Xeta cixdi"

        ws2[row[0].coordinate] = str(new_data)
        ws2[row[1].coordinate] = request_delivery_status
        ws2[row[2].coordinate] = response.text
        wb2.save(output_file_name)
        print(response ,'saved' , row_count)

        # Her 100 luy rowu kecende ayri faylda save elesin 

        # if row_count % 100 == 0:
        #     print(row_count)
        #     new_file_name = f"results[{row_count-100}-{row_count}]"
        #     wb2.save(new_file_name)
        
        row_count+=1
        time.sleep(5)    
        

    save_setting(settings)
    
    wb2.save(output_file_name)
    sys.exit('Finished')


if __name__  == '__main__':
    

    try:
        main()
    except KeyboardInterrupt:
        wb2.save(output_file_name)
        wb.save()
        save_setting(settings)
        sys.exit('Results and setting saved when keyboard interruption')





    





